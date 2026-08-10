import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from pulse_helpers import (
    check_for_blanks,
    create_results_workbook,
    is_my_workspace_report,
    make_report_options,
    parse_skip_visuals,
    should_skip_visual,
)


class BlankCheckTests(unittest.TestCase):
    def test_detects_blank_column_with_one_data_row(self):
        self.assertEqual(
            check_for_blanks("Category,Value\r\nExample,", "table"),
            ["Column 'Value' is blank"],
        )

    def test_warns_for_header_only_export(self):
        self.assertEqual(
            check_for_blanks("Category,Value\r\n", "table"),
            ["No data rows were exported"],
        )

    def test_ignores_slicers(self):
        self.assertIsNone(check_for_blanks("Choice\r\n", "slicer"))


class SelectionAndSkipTests(unittest.TestCase):
    def test_identifies_my_workspace_reports(self):
        self.assertTrue(
            is_my_workspace_report(
                {"webUrl": "https://app.powerbi.com/groups/me/reports/report-id"}
            )
        )
        self.assertFalse(
            is_my_workspace_report(
                {"webUrl": "https://app.powerbi.com/groups/shared-id/reports/report-id"}
            )
        )
        self.assertFalse(
            is_my_workspace_report(
                {"webUrl": "https://app.powerbi.com/groups/me/reports/report-id", "appId": "app-id"}
            )
        )

    def test_duplicate_report_labels_remain_selectable(self):
        reports = [
            {"workspace_name": "Workspace", "report_name": "Report"},
            {"workspace_name": "Workspace", "report_name": "Report"},
        ]
        options = make_report_options(reports)
        self.assertEqual(len(options), 2)
        self.assertNotEqual(options[0][0], options[1][0])
        self.assertEqual({value for _, value in options}, {0, 1})

    def test_parses_and_matches_generic_skip_rule(self):
        raw = json.dumps([{"report": "R", "page": "P", "visual": "V"}])
        rules = parse_skip_visuals(raw)
        self.assertTrue(should_skip_visual("R", "P", "V", rules))
        self.assertFalse(should_skip_visual("R", "P", "Other", rules))


class WorkbookTests(unittest.TestCase):
    def test_slow_only_run_creates_workbook(self):
        results = [
            {
                "status": "success",
                "criticality": 3,
                "report_name": "Synthetic report",
                "page_name": "Synthetic page",
                "visual_title": "Synthetic visual",
                "visual_type": "table",
                "error_message": "No error",
                "warnings": None,
                "export_time_seconds": 3.0,
            }
        ]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            created = create_results_workbook(
                results,
                [],
                path,
                slow_threshold_seconds=2.0,
                highlight_threshold_seconds=5.0,
            )
            self.assertEqual(created, path)
            self.assertEqual(load_workbook(path).sheetnames, ["Slow_Exports"])

    def test_slow_and_skipped_sheets_coexist(self):
        results = [{"status": "success", "export_time_seconds": 6.0}]
        skipped = [{"report_name": "R", "page_name": "P", "visual_title": "V"}]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            create_results_workbook(
                results,
                skipped,
                path,
                slow_threshold_seconds=2.0,
                highlight_threshold_seconds=5.0,
            )
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames, ["Slow_Exports", "Skipped_Visuals"])
            self.assertEqual(workbook["Slow_Exports"]["A2"].fill.fill_type, "solid")


if __name__ == "__main__":
    unittest.main()
