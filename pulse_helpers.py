"""Pure, locally testable helpers for the PULSE notebook."""

from __future__ import annotations

import csv
import io
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def make_workspace_options(workspaces: Sequence[Mapping[str, object]]) -> list[tuple[str, int]]:
    """Return unique widget labels paired with indexes into ``workspaces``."""
    names = [str(workspace.get("name") or "Unnamed workspace") for workspace in workspaces]
    return _make_unique_indexed_options(names)


def make_report_options(reports: Sequence[Mapping[str, object]]) -> list[tuple[str, int]]:
    """Return unique workspace/report labels paired with indexes into ``reports``."""
    labels = [
        f"{report.get('workspace_name') or 'Unnamed workspace'} - "
        f"{report.get('report_name') or 'Unnamed report'}"
        for report in reports
    ]
    return _make_unique_indexed_options(labels)


def is_my_workspace_report(report: Mapping[str, object]) -> bool:
    """Return whether an ungrouped API result belongs to My Workspace itself."""
    web_url = str(report.get("webUrl") or "").casefold()
    return not report.get("appId") and (
        "/groups/" not in web_url or "/groups/me/" in web_url
    )


def _make_unique_indexed_options(labels: Sequence[str]) -> list[tuple[str, int]]:
    totals = Counter(labels)
    seen: defaultdict[str, int] = defaultdict(int)
    options = []
    for index, label in enumerate(labels):
        seen[label] += 1
        suffix = f" [{seen[label]} of {totals[label]}]" if totals[label] > 1 else ""
        options.append((f"{label}{suffix}", index))
    return sorted(options, key=lambda option: option[0].casefold())


def parse_skip_visuals(raw_value: str | None) -> list[tuple[str, str, str]]:
    """Parse generic report/page/visual skip rules from a JSON environment value."""
    if not raw_value or not raw_value.strip():
        return []

    parsed = json.loads(raw_value)
    if not isinstance(parsed, list):
        raise ValueError("PULSE_SKIP_VISUALS_JSON must be a JSON list")

    rules: list[tuple[str, str, str]] = []
    for index, item in enumerate(parsed):
        if not isinstance(item, dict):
            raise ValueError(f"Skip rule {index + 1} must be a JSON object")
        values = tuple(item.get(key) for key in ("report", "page", "visual"))
        if not all(isinstance(value, str) and value.strip() for value in values):
            raise ValueError(
                f"Skip rule {index + 1} requires non-empty report, page, and visual strings"
            )
        rules.append(tuple(value.strip() for value in values))
    return rules


def should_skip_visual(
    report_name: str,
    page_name: str,
    visual_title: str,
    skip_rules: Iterable[tuple[str, str, str]],
) -> bool:
    """Return whether a visual exactly matches a configured skip rule."""
    candidate = tuple(value.strip() for value in (report_name, page_name, visual_title))
    return any(candidate == tuple(value.strip() for value in rule) for rule in skip_rules)


def check_for_blanks(exported_data: object, visual_type: str) -> list[str] | None:
    """Find columns whose exported values are entirely blank.

    Slicers and non-text results are intentionally excluded. A header-only export is
    surfaced as a warning because it is a useful potential blank-data signal.
    """
    if visual_type.casefold() == "slicer" or not isinstance(exported_data, str):
        return None
    if not exported_data.strip():
        return ["No data was exported"]

    rows = list(csv.reader(io.StringIO(exported_data)))
    if not rows or not any(header.strip() for header in rows[0]):
        return ["No data was exported"]

    headers = rows[0]
    data_rows = rows[1:]
    if not data_rows:
        return ["No data rows were exported"]

    warnings = []
    for index, header in enumerate(headers):
        values = [row[index].strip() if index < len(row) else "" for row in data_rows]
        if all(value == "" for value in values):
            warnings.append(f"Column '{header}' is blank")
    return warnings or None


def create_results_workbook(
    results: Sequence[Mapping[str, object]],
    skipped_visuals: Sequence[Mapping[str, object]],
    output_path: str | Path,
    *,
    slow_threshold_seconds: float,
    highlight_threshold_seconds: float,
) -> Path | None:
    """Write issue, slow-export, and skip sheets when any result warrants output."""
    issues = [dict(result) for result in results if result.get("status") != "success"]
    issues.sort(key=lambda result: result.get("criticality", 3))
    slow_exports = [
        dict(result)
        for result in results
        if float(result.get("export_time_seconds") or 0) > slow_threshold_seconds
    ]

    if not issues and not slow_exports and not skipped_visuals:
        return None

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    def normalize(rows: Sequence[Mapping[str, object]]) -> list[dict[str, object]]:
        normalized = []
        for row in rows:
            item = dict(row)
            if isinstance(item.get("warnings"), list):
                item["warnings"] = ", ".join(str(value) for value in item["warnings"])
            normalized.append(item)
        return normalized

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if issues:
            pd.DataFrame(normalize(issues)).to_excel(
                writer, sheet_name="Errors_Warnings", index=False
            )
        if slow_exports:
            pd.DataFrame(normalize(slow_exports)).to_excel(
                writer, sheet_name="Slow_Exports", index=False
            )
        if skipped_visuals:
            pd.DataFrame(skipped_visuals).to_excel(
                writer, sheet_name="Skipped_Visuals", index=False
            )

    if slow_exports:
        workbook = load_workbook(output_path)
        sheet = workbook["Slow_Exports"]
        headers = [cell.value for cell in sheet[1]]
        duration_column = headers.index("export_time_seconds")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        for row in sheet.iter_rows(min_row=2):
            duration = row[duration_column].value
            if duration is not None and float(duration) > highlight_threshold_seconds:
                for cell in row:
                    cell.fill = red_fill
        workbook.save(output_path)

    return output_path
